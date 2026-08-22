"""The workbook export, and the one property it shares with the Markdown.

**Two documents describing one run must not disagree.** A card that says
94.2% in one file and 94% in the other is worse than having a single
format: somebody forwards whichever copy suits them, and the two get
quoted against each other. So `decision_export` decides every value and
both renderers read it from there — and the test that matters most here
is the one checking the two files actually carry the same strings.

The rest is what a spreadsheet gets wrong that a page does not: a blank
cell sorts to the top and sums as zero, and a workbook is the format
most likely to be pulled apart one sheet at a time.
"""

from __future__ import annotations

import io

import pytest

from planbench_api.decision_markdown import render_decision_markdown
from planbench_api.decision_xlsx import decision_workbook_filename, render_decision_xlsx

ILLEGAL_SHEET_CHARACTERS = set("[]:*?/\\")

#: The sheets whose cells are the strings the Markdown prints, character
#: for character. The rest store raw values with a number format and are
#: checked against the screen instead — see
#: `test_neither_invents_a_number_the_other_lacks`.
MIRRORS_THE_MARKDOWN = (
    "Provenance",
    "Sample",
    "Gates",
    "Outcome by candidate",
    "Decision Card",
    "Episodes",
    "Human record",
)


class Run:
    """A stored run, in the shape both renderers read."""

    def __init__(self, report: dict, card: dict | None = None) -> None:
        self.id = "run_x"
        self.task_profile_id = "warehouse_a_v1"
        self.experiment_scope = "global_planner_selection"
        self.contracts_version = "6.9.0"
        self.created_at = "2026-08-20T00:00:00Z"
        self.report = report
        self.card = card
        self.review_state = "unreviewed"
        self.reviewed_by = None
        self.reviewed_at = None
        self.config_state = "pending"
        self.config_decided_by = None
        self.config_decided_at = None


def candidate(**overrides) -> dict:
    base = {
        "stack_label": "astar+dwa",
        "local_controller_config": "dwa_coarse",
        "local_observation_class": "lidar_only",
        "n_distinct_episodes": 30,
        "success_rate": 1.0,
        "pooled_p99_latency_ms": 7.3479,
        "replan_count": 30,
        "cleared_gates": True,
        "blocking_gates": [],
        "worst_clearance_m": 0.494,
        "median_travel_time_s": 22.8,
        "decision_utility": 0.8774,
        "objectives": {"U_R": 1.0, "U_S": 1.0, "U_E": 0.568, "U_C": 0.958},
        "recommendation_eligible": True,
        # A passing episode and a failing one, so the episode sheet is
        # exercised by the default fixture rather than only where a test
        # remembers to add rows.
        "episodes": [
            {
                "episode_context_id": "ep00", "success": True, "failure_reason": None,
                "collision_count": 0, "min_clearance": 0.494, "travel_time_s": 22.8,
                "p99_latency_ms": 11.1, "replan_count": 1,
                "episode_decision_utility": 0.88,
            },
            {
                "episode_context_id": "ep01", "success": False, "failure_reason": "timeout",
                "collision_count": 0, "min_clearance": 0.113, "travel_time_s": 60.0,
                "p99_latency_ms": 2098.4, "replan_count": 17,
                "episode_decision_utility": 0.31,
            },
        ],
    }
    base.update(overrides)
    return base


def report(**overrides) -> dict:
    base = {
        "identity": {"git_sha": "abc123", "anchor_config_version": "v1.2"},
        "sample": {"n_episodes": 30, "n_min_required": 30},
        "candidates": [candidate(), candidate(local_controller_config="dwa_balanced")],
    }
    base.update(overrides)
    return base


CARD = {
    "recommended": {"stack": "astar+dwa", "candidate_id": "e1251e42a20b"},
    "alternative": None,
    "status": "CLEAR_RECOMMENDATION",
    "contracts_version": "6.9.0",
    "recommendation_scope": "MISSION_LEVEL",
    # The card's own score, which the summary reads. Carried by the
    # default fixture rather than added per test: a card without it is
    # the unusual case, and leaving it out here made every assertion
    # about the score have to build its own card first.
    "decision_utility": 0.8774,
    "evidence": {
        "weight_stability_margin": 1.0,
        "anchor_stability": "unchanged",
        "robustness_margin": None,
    },
}


def book(run: Run):
    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(render_decision_xlsx(run)))


def cells(sheet) -> list[str]:
    return [
        str(value)
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]


class TestTheTwoExportsAgree:
    def test_every_gate_number_appears_in_both(self) -> None:
        """The property this whole split exists for.

        Not "both files mention the candidate" — the actual formatted
        strings. `7.35 ms` in one and `7.3479809999` in the other would
        be one measurement rendered two ways, and a reader holding both
        would have no way to tell.
        """
        run = Run(report(), CARD)
        markdown = render_decision_markdown(run)
        sheet = cells(book(run)["Gates"])
        for value in ("astar+dwa", "dwa_coarse", "lidar_only", "100.0%", "7.35 ms", "passed"):
            assert value in sheet, value
            assert value in markdown, value

    def test_the_scope_caveat_is_in_both(self) -> None:
        """A workbook is the format most likely to be pulled apart, so a
        caveat that travels only with the Markdown is one that stops
        travelling the moment somebody prefers spreadsheets."""
        run = Run(report(), CARD)
        assert "HĐ-1.4" in render_decision_markdown(run)
        assert any("HĐ-1.4" in value for value in cells(book(run)["Decision Card"]))

    def test_neither_invents_a_number_the_other_lacks(self) -> None:
        """Crude but real: every numeric-looking cell on the sheets that
        mirror the Markdown has to be findable in the Markdown.

        **Scoped to those sheets on purpose.** The numeric sheets store
        raw floats shown to a fixed number of decimals, while the
        Markdown prints three significant digits — the same measurement,
        a different number of digits, and no format string reproduces
        `.3g`. Asserting character equality across that line would force
        one of the two to give up what it exists for: the Markdown its
        frozen output, or the new sheets their sortable columns.
        """
        run = Run(report(), CARD)
        markdown = render_decision_markdown(run)
        workbook = book(run)
        for name in MIRRORS_THE_MARKDOWN:
            for value in cells(workbook[name]):
                if any(character.isdigit() for character in value) and len(value) < 40:
                    assert value in markdown, f"{name}: {value!r} is not in the Markdown"


class TestWhatASpreadsheetGetsWrong:
    def test_a_missing_value_says_so_rather_than_leaving_a_blank(self) -> None:
        """An empty cell sorts to the top, sums as zero, and averages as
        though it had been measured."""
        run = Run(report(), CARD)
        assert "not measured" in cells(book(run)["Decision Card"])

    def test_a_run_with_no_card_still_exports(self) -> None:
        """Fewer than two candidates through the gates means no ΔU
        (HĐ-7), and the gate table is then the whole deliverable."""
        run = Run(report(why_no_card="only one candidate cleared"), None)
        workbook = book(run)
        assert "Gates" in workbook.sheetnames
        card_sheet = cells(workbook["Decision Card"])
        assert any("No Decision Card" in value for value in card_sheet)
        assert any("only one candidate cleared" in value for value in card_sheet)

    def test_the_unlike_inputs_finding_lands_on_the_gate_sheet(self) -> None:
        """Beside the numbers it qualifies, not in a preamble somebody
        leaves behind when they copy one sheet into a slide."""
        run = Run(
            report(
                candidates=[candidate(), candidate(local_observation_class="full_static_map")]
            ),
            CARD,
        )
        assert any("shown different things" in value for value in cells(book(run)["Gates"]))

    def test_a_retired_candidate_is_named_on_the_gate_sheet(self) -> None:
        run = Run(
            report(
                candidates=[
                    candidate(),
                    candidate(
                        stack_label="rrtstar+dwa",
                        stopped_early={
                            "episodes_run": 12,
                            "episodes_planned": 30,
                            "gate": "G3",
                            "rule": "arithmetically doomed",
                        },
                    ),
                ]
            ),
            CARD,
        )
        found = cells(book(run)["Gates"])
        assert any("rrtstar+dwa" in value and "12 of 30" in value for value in found)


class TestTheFileItself:
    def test_it_is_a_workbook_a_reader_can_open(self) -> None:
        workbook = book(Run(report(), CARD))
        assert workbook.sheetnames == ["Summary", *MIRRORS_THE_MARKDOWN]

    def test_the_summary_comes_first(self) -> None:
        """A reader opening the file to check one number should not have
        to know which of eight tabs it lives on."""
        assert book(Run(report(), CARD)).sheetnames[0] == "Summary"

    def test_no_sheet_name_breaks_excel(self) -> None:
        """Past 31 characters, or carrying one of the reserved
        characters, Excel refuses the whole file — and the refusal
        arrives at the reader, not here."""
        for name in book(Run(report(), CARD)).sheetnames:
            assert len(name) <= 31
            assert not set(name) & ILLEGAL_SHEET_CHARACTERS

    def test_the_gate_header_stays_put_while_scrolling(self) -> None:
        assert book(Run(report(), CARD))["Gates"].freeze_panes == "A2"

class TestTheFilenameSaysWhatTheFileIsAbout:
    """`decision-8f3a1c.xlsx` is unambiguous and says nothing.

    A reader with six downloads in a folder needs the deployment and the
    pair being compared, and needs them without opening anything.
    """

    def test_it_names_the_project_the_comparison_and_when(self) -> None:
        run = Run(
            report(
                identity={"created_at": "2026-08-21T14:30:00+00:00"},
                candidates=[candidate(), candidate(stack_label="rrtstar+dwa")],
            ),
            CARD,
        )
        assert decision_workbook_filename(run) == (
            "warehouse_a_v1_astar-dwa-vs-rrtstar-dwa_2026-08-21_14-30.xlsx"
        )

    def test_two_exports_of_one_run_produce_one_name(self) -> None:
        """The timestamp is when the run happened, not when the button
        was pressed — otherwise two downloads of one run land as two
        files with no way to see they are the same document."""
        run = Run(report(), CARD)
        assert decision_workbook_filename(run) == decision_workbook_filename(run)

    def test_more_than_two_candidates_are_counted_rather_than_listed(self) -> None:
        """`a-vs-b-vs-c-vs-d` is a filename nobody reads to the end."""
        run = Run(report(candidates=[candidate(stack_label=f"s{n}") for n in range(4)]), CARD)
        assert "_4-candidates_" in decision_workbook_filename(run)

    def test_a_run_with_no_candidates_still_names_itself(self) -> None:
        run = Run(report(candidates=[]), None)
        assert decision_workbook_filename(run).startswith("warehouse_a_v1_no-candidates")

    def test_characters_a_filesystem_would_refuse_are_replaced(self) -> None:
        """`astar+dwa` is a stack name, not a filename."""
        run = Run(report(candidates=[candidate(stack_label="a*b/c:d")]), CARD)
        name = decision_workbook_filename(run)
        assert not set(name) & set('*/:\\?"<>|')

    def test_an_unparseable_timestamp_drops_the_segment_rather_than_inventing_one(self) -> None:
        """A name that states the wrong time is worse than one that
        states no time, because only the first is believed."""
        run = Run(
            report(
                identity={"created_at": "sometime last week"},
                candidates=[candidate()],
            ),
            CARD,
        )
        assert decision_workbook_filename(run) == "warehouse_a_v1_astar-dwa.xlsx"

    def test_it_stays_short_enough_to_save_on_windows(self) -> None:
        run = Run(report(candidates=[candidate(stack_label="x" * 200)]), CARD)
        assert len(decision_workbook_filename(run)) <= 125


class TestOverHttp:
    def test_the_route_serves_a_workbook(self, client, app, alice_headers) -> None:
        from planbench_api.decisions import StoredDecisionRun

        app.state.repos.decision_runs.create(
            StoredDecisionRun(
                id="run_xlsx",
                task_profile_id="warehouse_a_v1",
                artifact_kind="comparison",
                experiment_scope="global_planner_selection",
                contracts_version="6.9.0",
                created_at="2026-08-20T10:00:00Z",
                created_by=None,
                report=report(),
                card=None,
                manifest=None,
                recommended_candidate_id=None,
                status="unranked",
            )
        )
        response = client.get("/api/v1/decisions/run_xlsx/report.xlsx", headers=alice_headers)
        assert response.status_code == 200, response.text
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument"
        )
        disposition = response.headers["content-disposition"]
        assert "warehouse_a_v1_" in disposition
        assert "_2026-08-20_10-00.xlsx" in disposition
        # A real workbook, not an error page wearing the right header.
        assert response.content[:2] == b"PK"

    def test_an_unknown_run_is_a_404(self, client, alice_headers) -> None:
        response = client.get("/api/v1/decisions/nope/report.xlsx", headers=alice_headers)
        assert response.status_code == 404


@pytest.mark.parametrize("missing", ["report", "candidates"])
def test_a_thin_report_does_not_crash_the_export(missing: str) -> None:
    """A run stored before a field existed still has to export.

    The alternative is that the oldest runs — the ones somebody is
    digging up precisely because they are old — are the ones that
    cannot be shared.
    """
    if missing == "candidates":
        body = report()
        body["candidates"] = []
        run = Run(body, None)
    else:
        run = Run({}, None)
    assert render_decision_xlsx(run)[:2] == b"PK"


class TestWhatMakesItEvaluable:
    """The two additions that turned a result notice into a record.

    An, on the first version: "chỉ dùng để báo kết quả sơ bộ thôi, không
    đủ để đem ra đánh giá."
    """

    def test_the_margin_travels_with_its_interval(self) -> None:
        """ΔU alone turns "ahead, but not measurably" into a result."""
        run = Run(report(), {**CARD, "evidence": {**CARD["evidence"], "delta_u_mean": 0.039,
                                                  "ci95": [0.036, 0.042], "effect_size": 5.07,
                                                  "n_episodes": 30, "delta_u_vs_second": 0.036}})
        found = cells(book(run)["Decision Card"])
        assert any("ΔU mean" in value for value in found)
        assert any("interval" in value for value in found)
        assert "0.039" in found
        assert any(value.startswith("[0.036") for value in found)

    def test_every_episode_is_listed_with_how_it_failed(self) -> None:
        """`success_rate: 0.70` does not say *which* thirty per cent, nor
        whether they were collisions or timeouts — and those two ask for
        different work."""
        found = cells(book(Run(report(), CARD))["Episodes"])
        assert "ep00" in found and "ep01" in found
        assert "passed" in found
        assert "timeout" in found  # not a bare "failed"
        assert "0.113 m" in found

    def test_the_outcome_table_carries_what_the_page_compares_on(self) -> None:
        """Six of the ten metrics on the comparison grid never left the
        screen before."""
        found = cells(book(Run(report(), CARD))["Outcome by candidate"])
        for column in (
            "Utility /100", "Collisions", "Collision bound 95%", "No route found",
            "Worst clearance", "Median episode", "Memory estimate",
        ):
            assert column in found, column

    def test_eligibility_is_stated_rather_than_inferred(self) -> None:
        """A gate failure can leave no mark on the utility — collisions
        are excluded from U_S by contract — so "scored lower" and "was
        never in the running" cannot be told apart from the mark."""
        body = report()
        body["candidates"][0]["recommendation_eligible"] = False
        found = cells(book(Run(body, CARD))["Outcome by candidate"])
        assert "Eligible to recommend" in found
        assert "no" in found
        assert any("cannot be traded against speed" in value for value in found)


class TestTheSummarySheet:
    """The one sheet somebody reads if they read nothing else.

    Before it existed the answer to "what came out of this run?" was
    spread over three tabs: the identity on `Provenance`, the winner and
    the margin on `Decision Card`, and which stacks were even compared
    only on the tables underneath.
    """

    def sheet(self, run: Run):
        return book(run)["Summary"]

    def test_it_carries_what_a_reader_came_for(self) -> None:
        found = cells(self.sheet(Run(report(), CARD)))
        for label in (
            "Run",
            "Deployment",
            "Anchor config",
            "Contracts version",
            "Code version",
            "Winner",
            "Overall score",
            "Final recommendation",
        ):
            assert label in found, label

    def test_it_names_the_candidates_rather_than_calling_them_a_and_b(self) -> None:
        """"Algorithm A" makes a reader who opened the file a week later
        go and look up which one A was."""
        found = cells(self.sheet(Run(report(), CARD)))
        assert any("astar+dwa (dwa_coarse)" in value for value in found)

    def test_the_recommendation_carries_its_scope(self) -> None:
        """A recommendation that arrives without its limit is one
        somebody applies elsewhere (HĐ-1.4)."""
        found = " ".join(cells(self.sheet(Run(report(), CARD))))
        assert "MISSION_LEVEL" in found
        assert "HĐ-1.4" in found

    def test_a_run_with_no_card_still_gets_one(self) -> None:
        """"Nobody was ranked, here is why" is an answer a reader needs
        at the top, not four sheets in."""
        run = Run(report(why_no_card="only one candidate cleared"), None)
        found = " ".join(cells(self.sheet(run)))
        assert "Final recommendation" in found
        assert "ranked nobody" in found
        assert "only one candidate cleared" in found

    def test_it_names_the_preference_profile_when_the_run_recorded_one(self) -> None:
        """The weights the utility was computed under are part of what
        the score means, and two profiles rank the same candidates
        differently."""
        run = Run(report(manifest={"preference_profile": "benh_vien_gio_cao_diem"}), CARD)
        found = cells(self.sheet(run))
        assert "benh_vien_gio_cao_diem" in found

    def test_it_says_nothing_about_weights_a_run_did_not_record(self) -> None:
        """No manifest means no profile. Printing the default would be
        naming weights this card may not have been scored under."""
        assert "Preference profile" not in cells(self.sheet(Run(report(), CARD)))


class TestTheNumbersAreNumbers:
    """What a column of strings costs, and what buying it back costs.

    A spreadsheet whose figures are text cannot sort, sum or chart —
    which is most of the reason somebody asked for a spreadsheet. These
    sheets store the raw value and let Excel format it.
    """

    def cell(self, label: str, run: Run | None = None):
        sheet = book(run or Run(report(), CARD))["Summary"]
        for row in sheet.iter_rows():
            if row[0].value == label:
                return row[1]
        raise AssertionError(f"no row labelled {label!r}")

    def test_the_overall_score_is_a_number(self) -> None:
        cell = self.cell("Overall score")
        assert isinstance(cell.value, float)
        assert cell.number_format == "0.0000"

    def test_the_interval_is_two_numbers_rather_than_one_string(self) -> None:
        """A reader filtering for margins that clear zero needs the
        bound to be a number — which is the reading the interval exists
        to support."""
        card = {**CARD, "evidence": {**CARD["evidence"], "ci95": [0.032, 0.129]}}
        run = Run(report(), card)
        for label, expected in (("ΔU 95% lower bound", 0.032), ("ΔU 95% upper bound", 0.129)):
            cell = self.cell(label, run)
            assert cell.value == pytest.approx(expected)

    def test_a_missing_number_leaves_the_cell_empty_and_says_so_elsewhere(self) -> None:
        """Not 0, which sums and sorts as a measurement, and not the
        words "not measured", which would make the column text and take
        sorting away from every other row in it."""
        card = {**CARD, "decision_utility": None}
        cell = self.cell("Overall score", Run(report(), card))
        assert cell.value is None
        assert cell.number_format == "0.0000"

    def test_every_format_is_one_excel_accepts(self) -> None:
        from openpyxl import Workbook

        from planbench_api.decision_export import (
            COUNT,
            MEGABYTES,
            METRES,
            MILLISECONDS,
            PERCENT,
            SECONDS,
            UTILITY,
        )

        sheet = Workbook().active
        for index, unit in enumerate(
            (PERCENT, MILLISECONDS, SECONDS, METRES, MEGABYTES, COUNT, UTILITY), start=1
        ):
            cell = sheet.cell(row=index, column=1, value=1.0)
            cell.number_format = unit.excel_format
            assert cell.number_format == unit.excel_format

    def test_the_digits_match_what_the_screen_shows(self) -> None:
        """The fixed decimal counts are the comparison grid's, not new
        ones: the file and the page are the same comparison."""
        from planbench_api.decision_export import (
            COUNT,
            MEGABYTES,
            METRES,
            MILLISECONDS,
            PERCENT,
            SECONDS,
            quantity,
        )

        assert quantity(0.9667, PERCENT).display() == "96.7 %"
        assert quantity(7.3479, MILLISECONDS).display() == "7.35 ms"
        assert quantity(22.84, SECONDS).display() == "22.8 s"
        assert quantity(0.4942, METRES).display() == "0.494 m"
        assert quantity(412.53, MEGABYTES).display() == "412.5 MB"
        assert quantity(30, COUNT).display() == "30"

    def test_a_percentage_is_stored_raw_so_excels_own_format_scales_it(self) -> None:
        """`0.0%` multiplies by 100 itself. Storing 96.67 beside it would
        print 9667.0%."""
        from planbench_api.decision_export import PERCENT

        assert PERCENT.excel_format == "0.0%"
        assert PERCENT.display_scale == 100.0

    def test_a_difference_between_two_rates_is_points_not_per_cent(self) -> None:
        """Calling the gap between 70.0% and 72.0% "+2.0 %" says the gap
        is a proportion of a proportion."""
        from planbench_api.decision_export import PERCENT

        assert PERCENT.delta_symbol == "pp"

    def test_a_value_the_artifact_stored_as_text_does_not_break_the_export(self) -> None:
        """These come out of stored JSON. An older artifact with a
        string where a float belongs is a reason to leave the cell
        empty, not to fail the download."""
        from planbench_api.decision_export import METRES, quantity

        assert quantity("0.494", METRES).missing
        assert quantity(True, METRES).missing
