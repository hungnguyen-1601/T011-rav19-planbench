"""The two sheets the workbook gained, and the two claims that justify them.

**One: the file and the page are the same comparison.** The grid in
`apps/web/src/lib/candidateMetrics.ts` decides which metrics a reader
compares stacks on, which way is better, what unit each carries and how
close counts as level. A second opinion in Python would mean the person
looking at the screen and the person opening the download are reading
two different comparisons of one run — and neither would know. The
parity test below reads the TypeScript and refuses to let the two drift.

**Two: the weight column tells the truth about what was weighted.**
Weights attach to objectives, not to the metrics on the comparison
sheet: three of the ten rows carry one and seven do not, and that is the
contract (HĐ-6 keeps collisions out of U_S so they cannot be traded
against speed) rather than seven cells nobody filled in. The
`Objective Breakdown` sheet is where the arithmetic is visible, and the
test that its Contribution column sums to the card's own utility is the
single most load-bearing assertion in this file: it fails if the numbers
are text, if the weights came from the wrong profile, if an objective is
mapped to the wrong axis, or if a second scoring path has crept in.
"""

from __future__ import annotations

import io
import re
from pathlib import Path

import pytest

from planbench_api.decision_export import (
    COMPARISON_METRICS,
    TIE_TOLERANCE,
    comparison_rows,
    objective_rows,
    resolve_weights,
)
from planbench_api.decision_xlsx import render_decision_xlsx
from tests.api.golden_run import golden_run, report, unranked_run

GRID = (
    Path(__file__).resolve().parents[2]
    / "apps"
    / "web"
    / "src"
    / "lib"
    / "candidateMetrics.ts"
)


def book(run, locale="en"):
    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(render_decision_xlsx(run, locale)))


def rows_of(sheet):
    return list(sheet.iter_rows(values_only=True))


def header_of(sheet):
    return [value for value in rows_of(sheet)[0] if value is not None]


def find(sheet, label: str):
    for row in sheet.iter_rows():
        if row[0].value == label:
            return row
    raise AssertionError(f"no row labelled {label!r}")


class TestTheFileAndThePageAreTheSameComparison:
    """Parity with `candidateMetrics.ts`, read out of the source.

    The cheap half of unifying the two definitions. The expensive half —
    the report carrying this table so the page reads it back — is worth
    doing and is not worth blocking the export on; until then this test
    is what makes the duplication safe.
    """

    def grid_rows(self) -> list[tuple[str, str]]:
        source = GRID.read_text(encoding="utf8")
        body = source[source.index("export function comparisonRows") :]
        return re.findall(r'row\(\s*"(\w+)",\s*"(higher|lower|none)"', body)

    def test_the_same_metrics_in_the_same_order(self) -> None:
        assert [key for key, _ in self.grid_rows()] == [
            spec.key for spec in COMPARISON_METRICS
        ]

    def test_the_same_direction_on_every_row(self) -> None:
        """A row the page calls "lower is better" and the file calls
        "higher" would crown opposite winners on one measurement."""
        assert dict(self.grid_rows()) == {
            spec.key: spec.direction for spec in COMPARISON_METRICS
        }

    def test_the_same_tie_tolerance(self) -> None:
        """A row where the two are level must not name a winner on the
        page and a different one in the file."""
        source = GRID.read_text(encoding="utf8")
        found = re.search(r"const TIE_TOLERANCE = ([0-9.e-]+);", source)
        assert found, "candidateMetrics.ts no longer declares TIE_TOLERANCE"
        assert float(found.group(1)) == TIE_TOLERANCE

    def test_the_same_units_and_delta_units(self) -> None:
        """`pp` is the one that matters: the gap between two percentages
        is percentage points, and calling it `%` would say the gap was a
        proportion of a proportion."""
        source = GRID.read_text(encoding="utf8")
        declared = dict(re.findall(r'unit: "([^"]*)",\s*\n\s*deltaUnit: "([^"]+)"', source))
        # Only the rows whose delta unit differs from their unit are
        # declared in the grid; the rest inherit.
        assert declared.get("%") == "pp"
        percent = next(spec for spec in COMPARISON_METRICS if spec.key == "successRate")
        assert percent.unit.symbol == "%"
        assert percent.unit.delta_symbol == "pp"


class TestTheComparisonSheet:
    def sheet(self, run=None, locale="en"):
        return book(run or golden_run(), locale)[
            "Detailed Comparison" if locale == "en" else "So sánh chi tiết"
        ]

    def test_it_replaced_the_per_candidate_table_rather_than_joining_it(self) -> None:
        """Both would print one table twice in one file and leave a
        reader deciding which to trust."""
        names = book(golden_run()).sheetnames
        assert "Detailed Comparison" in names
        assert "Outcome by candidate" not in names

    def test_ten_rows_in_the_order_the_page_uses(self) -> None:
        labels = [row[0].value for row in list(self.sheet().iter_rows())[1:11]]
        assert labels == [
            "Success rate",
            "Collisions observed",
            "Collision probability, 95% upper bound",
            "Episodes with no route found",
            "Worst clearance in the whole run",
            "Median episode duration",
            "Planner latency, pooled p99",
            "Memory estimate on the target board",
            "Distinct episodes",
            "Replans across the run",
        ]

    def test_the_columns_name_the_stacks_rather_than_a_and_b(self) -> None:
        header = header_of(self.sheet())
        assert "astar+dwa" in header
        assert "rrtstar+dwa" in header
        assert "Algorithm A" not in header

    def test_the_delta_unit_of_a_rate_is_points(self) -> None:
        row = find(self.sheet(), "Success rate")
        assert row[5].value == "pp"

    def test_a_rate_is_stored_raw_under_excels_own_percent_format(self) -> None:
        row = find(self.sheet(), "Success rate")
        assert row[2].value == pytest.approx(1.0)
        assert row[2].number_format == "0.0%"

    def test_the_limit_column_carries_the_deployments_declared_bar(self) -> None:
        """`7.35 ms` is meaningless without knowing the ceiling is 50."""
        assert find(self.sheet(), "Planner latency, pooled p99")[7].value == pytest.approx(50.0)
        assert find(self.sheet(), "Success rate")[7].value == pytest.approx(0.95)

    def test_only_three_rows_carry_a_weight(self) -> None:
        """Weights attach to objectives, not to metrics. Seven blanks
        here are the contract, not seven cells nobody filled in."""
        weighted = [
            row[0].value
            for row in list(self.sheet().iter_rows())[1:11]
            if row[8].value is not None
        ]
        assert weighted == [
            "Success rate",
            "Planner latency, pooled p99",
            "Memory estimate on the target board",
        ]

    def test_every_unweighted_row_says_why_in_its_note(self) -> None:
        for row in list(self.sheet().iter_rows())[1:11]:
            if row[8].value is None:
                assert row[9].value, row[0].value
                assert len(row[9].value) > 40, row[0].value

    def test_the_replans_row_has_no_winner_and_no_colour(self) -> None:
        """Replanning is already charged in travel time and in latency.
        Marking a winner would price it twice under a rule nobody wrote."""
        row = find(self.sheet(), "Replans across the run")
        assert row[6].value == "no direction"
        for cell in (row[2], row[3]):
            assert cell.fill.fgColor.rgb in (None, "00000000")

    def test_a_level_row_says_tie_rather_than_naming_somebody(self) -> None:
        sheet = self.sheet()
        assert find(sheet, "Distinct episodes")[6].value == "tie"

    def test_a_metric_only_one_side_recorded_is_not_a_comparison(self) -> None:
        """"Not leading" is not "behind". A candidate that recorded
        nothing did not lose — there was no comparison."""
        thin = golden_run()
        thin.report["candidates"][1]["worst_clearance_m"] = None
        row = find(self.sheet(thin), "Worst clearance in the whole run")
        assert row[6].value == "not measured"
        assert row[2].fill.fgColor.rgb in (None, "00000000")
        assert row[3].value is None

    def test_the_leader_is_tinted_and_the_trailer_is_too(self) -> None:
        row = find(self.sheet(), "Planner latency, pooled p99")
        assert row[2].fill.fgColor.rgb not in (None, "00000000")
        assert row[3].fill.fgColor.rgb not in (None, "00000000")
        assert row[2].fill.fgColor.rgb != row[3].fill.fgColor.rgb

    def test_a_value_past_the_declared_limit_is_marked(self) -> None:
        slow = golden_run()
        slow.report["candidates"][1]["pooled_p99_latency_ms"] = 91.0
        row = find(self.sheet(slow), "Planner latency, pooled p99")
        assert row[3].fill.fgColor.rgb not in (None, "00000000")

    def test_three_candidates_drop_the_delta_and_name_every_leader(self) -> None:
        """`B minus A` has no meaning across three, and `leaders` returns
        a set because two of three can be equally best."""
        three = golden_run()
        third = dict(three.report["candidates"][0])
        third["stack_label"] = "prm+dwa"
        three.report["candidates"] = [*three.report["candidates"], third]
        sheet = self.sheet(three)
        assert "Delta" not in header_of(sheet)
        winner = find(sheet, "Success rate")[5].value
        assert "astar+dwa" in winner and "prm+dwa" in winner

    def test_a_run_with_no_card_still_gets_the_comparison(self) -> None:
        assert "Detailed Comparison" in book(unranked_run()).sheetnames

    def test_the_sheet_exists_in_vietnamese_with_translated_notes(self) -> None:
        sheet = self.sheet(locale="vi")
        assert "Chỉ số" in header_of(sheet)
        assert any("HĐ-6" in str(row[9].value) for row in list(sheet.iter_rows())[1:11])


class TestTheObjectiveBreakdown:
    def sheet(self, run=None, locale="en"):
        return book(run or golden_run(), locale)[
            "Objective Breakdown" if locale == "en" else "Phân rã theo mục tiêu"
        ]

    def test_the_contributions_sum_to_the_cards_own_utility(self) -> None:
        """The most load-bearing assertion here. It fails if the numbers
        are text, if the weights came from the wrong profile, if an
        objective is mapped to the wrong axis, or if a second scoring
        path has crept in."""
        sheet = self.sheet()
        axes = [find(sheet, f"Objective {name}") for name in ("U_R", "U_S", "U_E", "U_C")]
        total = find(sheet, "Decision utility")
        for column, expected in ((5, total[5].value), (6, total[6].value)):
            assert sum(row[column].value for row in axes) == pytest.approx(expected, abs=1e-6)

    def test_the_weights_come_from_the_runs_own_profile(self) -> None:
        """`kho_ban_dem` weights reliability at 0.30 and safety at 0.10.
        A card scored under another profile exported with these would
        print contributions that are all wrong and a total that still
        looks plausible."""
        sheet = self.sheet()
        assert find(sheet, "Objective U_R")[1].value == pytest.approx(0.30)
        assert find(sheet, "Objective U_S")[1].value == pytest.approx(0.10)

    def test_two_runs_under_different_profiles_get_different_weights(self) -> None:
        other = golden_run()
        other.manifest = {**other.manifest, "preference_profile": "benh_vien_gio_cao_diem"}
        assert find(self.sheet(other), "Objective U_S")[1].value == pytest.approx(0.50)

    def test_a_perturbed_run_prints_no_weight_and_says_why(self) -> None:
        """The HĐ-11.5 sweep does not record its replacements. Printing
        the named profile's numbers would attribute the card to weights
        it was not scored under."""
        swept = golden_run()
        swept.manifest = {**swept.manifest, "preference_profile": "kho_ban_dem (perturbed)"}
        sheet = self.sheet(swept)
        assert find(sheet, "Objective U_R")[1].value is None
        assert any("perturbed" in str(cell) for row in rows_of(sheet) for cell in row if cell)

    def test_a_profile_the_table_lost_is_named_rather_than_guessed(self) -> None:
        gone = golden_run()
        gone.manifest = {**gone.manifest, "preference_profile": "retired_profile"}
        sheet = self.sheet(gone)
        assert find(sheet, "Objective U_R")[1].value is None
        assert any("retired_profile" in str(cell) for row in rows_of(sheet) for cell in row if cell)

    def test_a_run_with_no_card_has_no_sheet_at_all(self) -> None:
        """Nothing was weighted, so a page of empty weights would be
        about an arithmetic that did not happen."""
        assert "Objective Breakdown" not in book(unranked_run()).sheetnames

    def test_the_cost_split_is_shown_with_its_weights(self) -> None:
        sheet = self.sheet()
        latency = find(sheet, "  ↳ Planner latency, pooled p99")
        # kho_ban_dem weights cost at 0.35, and β1 gives latency 0.30 of it.
        assert latency[1].value == pytest.approx(0.35 * 0.30)

    def test_the_two_unmeasured_components_carry_a_weight_and_no_value(self) -> None:
        """Weighted, and their inputs are nowhere in the report. Saying
        both halves is the whole point of the row."""
        sheet = self.sheet()
        for label in ("  ↳ CPU time per mission", "  ↳ Engineering cost"):
            row = find(sheet, label)
            assert row[1].value is not None
            assert row[2].value is None
            assert "not recorded in the report" in row[7].value

    def test_the_sheet_exists_in_vietnamese(self) -> None:
        assert "Mục tiêu" in header_of(self.sheet(locale="vi"))


class TestWeightsAreNeverGuessed:
    def test_a_run_with_no_manifest_resolves_to_nothing(self) -> None:
        run = unranked_run()
        assert resolve_weights(run, run.report).known is False

    def test_the_resolution_reads_the_run_and_not_a_default(self) -> None:
        run = golden_run()
        assert resolve_weights(run, run.report).profile == "kho_ban_dem"


class TestItSurvivesAThinReport:
    def test_no_candidates_means_no_comparison_rather_than_a_crash(self) -> None:
        empty = report()
        empty["candidates"] = []
        assert comparison_rows(empty) == []

    def test_a_candidate_with_no_gates_leaves_those_cells_empty(self) -> None:
        thin = report()
        for candidate in thin["candidates"]:
            candidate.pop("gates")
        rows = {row.label: row for row in comparison_rows(thin)}
        assert all(value.missing for value in rows["Collisions observed"].values)
        assert rows["Collisions observed"].limit is None

    def test_objectives_need_a_card(self) -> None:
        run = unranked_run()
        assert objective_rows(run, run.report) is None
