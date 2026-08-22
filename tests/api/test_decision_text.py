"""The bilingual text layer, and the two ways it could quietly go wrong.

**Wrong one: English moves.** The Markdown export is a document people
keep. `test_decision_export_golden` freezes it; this file guards the
narrower claim that the *default* is still English, because a default
that drifted to the other language would pass a golden test that always
passed the locale explicitly.

**Wrong two: a hole reads as a translation.** A key with no Vietnamese
would either raise or fall back to English. Falling back is the
dangerous one: the reader gets a document that is Vietnamese except for
the sentences nobody got round to, and cannot tell which of those were
left in English on purpose. Every entry is checked for both languages
here so the hole is found at import time rather than in somebody's
download.
"""

from __future__ import annotations

import pytest

from planbench_api.decision_markdown import render_decision_markdown
from planbench_api.decision_text import (
    DEFAULT_LOCALE,
    LOCALES,
    TEXT,
    MissingTranslation,
    lines,
    text,
)
from planbench_api.decision_xlsx import render_decision_xlsx
from tests.api.golden_run import golden_run, unranked_run

#: Excel truncates past this and the whole file then fails to open.
SHEET_NAME_LIMIT = 31


class TestEveryStringHasEveryLanguage:
    def test_no_entry_is_missing_a_language(self) -> None:
        missing = [
            f"{key}:{locale}"
            for key, entry in TEXT.items()
            for locale in LOCALES
            if not entry.get(locale)
        ]
        assert missing == [], missing

    def test_no_entry_carries_a_language_nobody_asked_for(self) -> None:
        """A stray third locale is a typo in a key, not a feature."""
        stray = {
            f"{key}:{locale}"
            for key, entry in TEXT.items()
            for locale in entry
            if locale not in LOCALES
        }
        assert stray == set(), stray

    def test_the_placeholders_match_across_languages(self) -> None:
        """A translation that drops `{scope}` renders a caveat with a
        hole in it, and one that invents `{name}` raises at render time
        in front of the reader."""
        import re

        holes = re.compile(r"\{(\w+)\}")
        for key, entry in TEXT.items():
            fields = {locale: set(holes.findall(value)) for locale, value in entry.items()}
            assert len(set(map(frozenset, fields.values()))) == 1, (key, fields)

    def test_a_missing_key_raises_rather_than_returning_something(self) -> None:
        with pytest.raises(MissingTranslation):
            text("no.such.key")

    def test_a_missing_language_raises_rather_than_serving_english(self) -> None:
        """The whole reason this is not a `.get(locale, en)`."""
        TEXT["test.only.english"] = {"en": "only english"}  # type: ignore[typeddict-item]
        try:
            with pytest.raises(MissingTranslation):
                text("test.only.english", "vi")
        finally:
            del TEXT["test.only.english"]


class TestEnglishIsStillTheDefault:
    """A default that drifted would pass every test that names a locale."""

    def test_the_default_locale_is_english(self) -> None:
        assert DEFAULT_LOCALE == "en"

    @pytest.mark.parametrize("build", [golden_run, unranked_run])
    def test_markdown_with_no_locale_equals_markdown_with_english(self, build) -> None:
        assert render_decision_markdown(build()) == render_decision_markdown(build(), "en")

    @pytest.mark.parametrize("build", [golden_run, unranked_run])
    def test_the_workbook_with_no_locale_equals_the_workbook_with_english(self, build) -> None:
        assert render_decision_xlsx(build()) == render_decision_xlsx(build(), "en")


class TestVietnameseIsAWholeDocument:
    def test_the_markdown_headings_are_translated(self) -> None:
        rendered = render_decision_markdown(golden_run(), "vi")
        assert "## Nguồn gốc" in rendered
        assert "## Thẻ quyết định" in rendered
        # And the English headings are gone rather than sitting beside them.
        assert "## Provenance" not in rendered
        assert "## Decision Card" not in rendered

    def test_the_caveats_are_translated_and_not_dropped(self) -> None:
        """The caveats are the part that does the work. A Vietnamese
        document that kept the numbers and lost the conditions for
        reading them would be the worse half of a translation."""
        rendered = render_decision_markdown(golden_run(), "vi")
        assert "HĐ-1.4" in rendered
        assert "khuyến nghị này áp dụng cho" in rendered
        assert "khoảng tin cậy" in rendered
        assert "HĐ-14" in rendered

    def test_an_unranked_run_says_why_in_vietnamese_too(self) -> None:
        rendered = render_decision_markdown(unranked_run(), "vi")
        assert "## Không có Thẻ quyết định" in rendered
        assert "ΔU không tồn tại" in rendered

    def test_not_measured_is_translated_rather_than_left_in_english(self) -> None:
        """HĐ-12 makes null a finding, so the word carrying that finding
        has to be one the reader reads."""
        thin = golden_run()
        thin.report["candidates"][0]["worst_clearance_m"] = None
        rendered = render_decision_markdown(thin, "vi")
        assert "chưa đo" in rendered
        assert "not measured" not in rendered


class TestTheWorkbookSurvivesTheLongerLanguage:
    @pytest.mark.parametrize("locale", LOCALES)
    @pytest.mark.parametrize("build", [golden_run, unranked_run])
    def test_no_sheet_name_breaks_excel_in_either_language(self, locale, build) -> None:
        """Vietnamese runs longer than English, and Excel refuses the
        whole file over a 32-character sheet name — a failure that
        arrives at the reader, not here."""
        import io

        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(render_decision_xlsx(build(), locale)))
        for name in workbook.sheetnames:
            assert len(name) <= SHEET_NAME_LIMIT, name
            assert not set(name) & set("[]:*?/\\"), name

    def test_the_two_languages_produce_the_same_sheet_count(self) -> None:
        """Same document, different words. A language that lost a sheet
        would mean a branch reached through a translated string."""
        import io

        import openpyxl

        counts = {
            locale: len(
                openpyxl.load_workbook(
                    io.BytesIO(render_decision_xlsx(golden_run(), locale))
                ).sheetnames
            )
            for locale in LOCALES
        }
        assert len(set(counts.values())) == 1, counts


class TestTheLineBreaksBelongToTheText:
    def test_a_paragraph_keeps_its_wrapping_in_both_languages(self) -> None:
        """Markdown quotes these line by line and the spreadsheet joins
        them with spaces, so the shape travels with the words rather
        than being re-derived per format."""
        for locale in LOCALES:
            assert len(lines("prose.gates", locale)) >= 2
            assert len(lines("caveat.mixed.body", locale)) >= 2
