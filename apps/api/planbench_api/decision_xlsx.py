"""A selection run as one workbook.

The Markdown export goes in a ticket; this one goes to somebody who
works in a spreadsheet. Same content, same words, different container —
`decision_export` decides every value, and this module only decides
where it sits.

**What a spreadsheet gets wrong that a page does not.** Two of the
export's structural properties are sharper here:

- *Null renders as "not measured", never as a blank.* On a page a blank
  cell is merely unhelpful. In a spreadsheet it sorts to the top, sums
  as zero, and averages as though it were measured.
- *The caveats travel.* A workbook is the format most likely to be
  pulled apart — one sheet copied into a slide, one column pasted into
  an email — so each caveat sits on the sheet holding the numbers it
  qualifies rather than in a preamble somebody will leave behind.

**Values are written as the strings the Markdown export shows**, not as
raw floats. It costs sorting and summing, and it buys the one property
worth more: the two documents cannot quote different numbers for the
same run. `7.35 ms` here and `7.3479809999` in the other file would be
the same value and a reader would have no way to know it.
"""

from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Any

from planbench_api.decision_export import (
    COUNT,
    DEFAULT_LOCALE,
    UTILITY,
    WEIGHT,
    Locale,
    Quantity,
    as_text,
    card_rows,
    comparison_rows,
    decision_evidence_rows,
    eligibility_row,
    environment_warning,
    episode_columns,
    episode_rows,
    gate_columns,
    gate_rows,
    human_rows,
    mixed_observation,
    no_card_reason,
    objective_rows,
    provenance_rows,
    resolve_weights,
    retired_candidates,
    sample_rows,
    scope_of,
    sensitivity_rows,
    summary_rows,
)
from planbench_api.decision_text import text

__all__ = ["decision_workbook_filename", "render_decision_xlsx"]

#: Excel refuses these in a sheet name and truncates past 31 characters.
#: Enforced rather than assumed: a name that trips either rule makes the
#: whole file unopenable, and the failure arrives at the reader, not here.
_ILLEGAL_IN_SHEET_NAME = re.compile(r"[\[\]:*?/\\]")
_SHEET_NAME_LIMIT = 31

#: Anything outside this becomes a hyphen. Deliberately narrower than
#: what any one filesystem forbids: the file is downloaded by a browser,
#: saved on whatever the reader runs, and mailed on from there, so the
#: safe set is the intersection rather than the union.
_UNSAFE_IN_FILENAME = re.compile(r"[^A-Za-z0-9._-]+")

#: Long enough for two stack names and a deployment, short enough that
#: the whole path stays inside the Windows limit after it lands in a
#: nested Downloads folder.
_FILENAME_STEM_LIMIT = 120


def _slug(value: Any) -> str:
    """A filename-safe fragment, or the empty string when there is none."""
    return _UNSAFE_IN_FILENAME.sub("-", str(value or "")).strip("-")


def _comparison_slug(report: dict[str, Any]) -> str:
    """What was compared, named in the filename.

    Two candidates are named; more are counted. `a-vs-b-vs-c-vs-d` is a
    filename nobody reads to the end, and the count is the part that
    tells a reader which download this is.
    """
    labels = [_slug(entry.get("stack_label")) for entry in report.get("candidates") or []]
    labels = [label for label in labels if label]
    if not labels:
        return "no-candidates"
    if len(labels) > 2:
        return f"{len(labels)}-candidates"
    return "-vs-".join(labels)


def _when(run: Any, report: dict[str, Any]) -> str:
    """When the run happened, or nothing at all.

    **Not the moment the button was pressed.** Two exports of one run
    have to produce one filename, or a reader with both in a folder has
    two files and no way to see they are the same document.

    An unparseable timestamp drops the segment rather than substituting
    now(): a name that states the wrong time is worse than one that
    states no time, because only the first is believed.
    """
    identity = report.get("identity") or {}
    raw = identity.get("created_at") or getattr(run, "created_at", None)
    if not raw:
        return ""
    try:
        stamp = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except ValueError:
        return ""
    return stamp.strftime("%Y-%m-%d_%H-%M")


def decision_workbook_filename(run: Any) -> str:
    """``<project>_<comparison>_<YYYY-MM-DD_HH-mm>.xlsx``.

    Named for what the file is about rather than for the id it was
    fetched by. `decision-8f3a1c.xlsx` is unambiguous and says nothing;
    a reader with six of them in a Downloads folder needs the deployment
    and the pair, and needs them without opening anything.
    """
    report: dict[str, Any] = getattr(run, "report", None) or {}
    parts = [
        _slug(getattr(run, "task_profile_id", None)) or "decision",
        _comparison_slug(report),
    ]
    when = _when(run, report)
    if when:
        parts.append(when)
    return f"{'_'.join(parts)[:_FILENAME_STEM_LIMIT]}.xlsx"


def _sheet_name(title: str) -> str:
    return _ILLEGAL_IN_SHEET_NAME.sub("-", title)[:_SHEET_NAME_LIMIT]


def render_decision_xlsx(run: Any, locale: Locale = DEFAULT_LOCALE) -> bytes:
    """The whole run as a workbook, one sheet per section."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    report: dict[str, Any] = run.report or {}
    workbook = Workbook()
    workbook.remove(workbook.active)

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    # Painted from the winner this module already computed, not from a
    # threshold rule Excel would evaluate on its own. A second rule beside
    # the `Winner` column could disagree with it, and a reader would have
    # two verdicts on one row. `breach` is the exception and is genuinely
    # a threshold: the deployment declared the number, so it is a fact
    # about the value rather than about the comparison.
    ahead_fill = PatternFill("solid", fgColor="DCF3E3")
    behind_fill = PatternFill("solid", fgColor="FBE3E4")
    breach_fill = PatternFill("solid", fgColor="FDEBD0")

    def sheet(key: str):
        """A sheet named from the text table.

        Named by key rather than by literal because Vietnamese runs
        longer than English and Excel truncates a sheet name at 31
        characters — a limit that has to be checked in whichever
        language the reader asked for, not only the default.
        """
        return workbook.create_sheet(_sheet_name(text(key, locale)))

    def write_value(target, row: int, column: int, value: str | Quantity):
        """One cell, as a number where the value is one.

        A :class:`Quantity` reaches the sheet as a float with a format,
        not as the string a reader would see. That is what makes the
        column sortable and summable — and it is why a missing value
        leaves the cell genuinely **empty** rather than carrying the
        words "not measured": text in a numeric column takes sorting
        away from every other row in it, and 0 would be read as a
        measurement.
        """
        cell = target.cell(row=row, column=column)
        if isinstance(value, Quantity):
            if not value.missing:
                cell.value = value.value
            cell.number_format = value.unit.excel_format
        else:
            cell.value = value
        return cell

    def write_pairs(target, rows, start: int) -> int:
        for offset, (label, value) in enumerate(rows):
            target.cell(row=start + offset, column=1, value=label).font = bold
            write_value(target, start + offset, 2, value)
        return start + len(rows)

    def write_caveat(target, label: str, body: str, at: int) -> int:
        """A caveat, on the sheet holding what it qualifies.

        Labelled rather than merged into the surrounding rows: a reader
        copying a range out should take the warning with the numbers or
        leave both.

        The parameter is `body` and not `text`: this module now imports a
        function by that name, and a parameter shadowing it would leave
        the next line added inside here calling a string.
        """
        target.cell(row=at, column=1, value=label).font = bold
        cell = target.cell(row=at, column=2, value=body)
        cell.alignment = wrap
        return at + 1

    # --- Summary ----------------------------------------------------------
    #
    # First, and the only sheet that answers "what came out of this?"
    # without the reader assembling it from three others: the run's
    # identity was on `Provenance`, the winner and the margin on
    # `Decision Card`, and which stacks were compared only on the tables
    # underneath. A reader who opens the file to check one number should
    # not have to know which tab it lives on.
    summary = sheet("heading.summary")
    cursor = write_pairs(summary, summary_rows(run, report, locale), 1)
    write_caveat(
        summary,
        text("heading.precision", locale),
        text("prose.summary_precision", locale),
        cursor + 1,
    )
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 82

    # --- Provenance -------------------------------------------------------
    provenance = sheet("heading.provenance")
    heading = provenance.cell(
        row=1,
        column=1,
        value=text(
            "heading.document", locale, profile=as_text(run.task_profile_id, locale)
        ),
    )
    heading.font = bold
    cursor = write_pairs(provenance, provenance_rows(run, report, locale), 3)
    warning = environment_warning(report, locale)
    if warning:
        write_caveat(
            provenance,
            text("heading.measurement_environment", locale),
            warning,
            cursor + 1,
        )
    provenance.column_dimensions["A"].width = 26
    provenance.column_dimensions["B"].width = 82

    # --- Sample -----------------------------------------------------------
    sample = sheet("heading.sample")
    write_pairs(sample, sample_rows(report, locale), 1)
    sample.column_dimensions["A"].width = 30
    sample.column_dimensions["B"].width = 20

    # --- Gates ------------------------------------------------------------
    candidates = report.get("candidates") or []
    if candidates:
        gates = sheet("heading.gates")
        for column, label in enumerate(gate_columns(locale), start=1):
            gates.cell(row=1, column=column, value=label).font = bold
        for index, row in enumerate(gate_rows(report, locale), start=2):
            for column, value in enumerate(row, start=1):
                gates.cell(row=index, column=column, value=value)
        # The header stays put while a reader scrolls a long field.
        gates.freeze_panes = "A2"
        for column, width in zip("ABCDEFGH", (22, 16, 18, 18, 12, 14, 12, 34), strict=False):
            gates.column_dimensions[column].width = width

        cursor = len(candidates) + 3
        mixed = mixed_observation(candidates, locale)
        if mixed:
            cursor = write_caveat(
                gates, text("heading.unlike_inputs", locale), mixed.sentence(), cursor
            )
        retired = retired_candidates(candidates, locale)
        if retired:
            gates.cell(
                row=cursor + 1, column=1, value=text("heading.retired_early", locale)
            ).font = bold
            for offset, (label, detail) in enumerate(retired):
                gates.cell(row=cursor + 1 + offset, column=2, value=f"{label} — {detail}")

    # --- What the sweep concluded about each candidate --------------------
    def write_table(target, columns, rows, widths) -> None:  # noqa: ANN001, ANN202
        for column, label in enumerate(columns, start=1):
            target.cell(row=1, column=column, value=label).font = bold
        for index, row in enumerate(rows, start=2):
            for column, value in enumerate(row, start=1):
                target.cell(row=index, column=column, value=value)
        # The header stays put while a reader scrolls a long field.
        target.freeze_panes = "A2"
        for offset, width in enumerate(widths):
            target.column_dimensions[get_column_letter(offset + 1)].width = width

    # --- Detailed comparison ----------------------------------------------
    #
    # **This replaced `Outcome by candidate`, it does not sit beside it.**
    # That sheet was the same ten figures laid out one candidate per row,
    # and every column it had is here or on `Objective Breakdown`. Keeping
    # both would print one table twice in one file and leave a reader
    # deciding which to trust.
    #
    # Transposed because a spreadsheet reader compares *down* a column: a
    # metric per row is what makes room for the delta, the winner and the
    # deployment's own limit beside the two figures being compared.
    weights = resolve_weights(run, report)
    comparison = comparison_rows(report, locale, weights)
    if comparison:
        labels = [
            as_text(entry.get("stack_label"), locale) for entry in report.get("candidates") or []
        ]
        # The stacks by name. "Algorithm A" makes a reader who opened the
        # file a week later go and look up which one A was.
        header = [
            text("column.compare.metric", locale),
            text("column.compare.unit", locale),
            *labels,
        ]
        has_delta = any(row.delta is not None for row in comparison)
        if has_delta:
            header += [
                text("column.compare.delta", locale),
                text("column.compare.delta_unit", locale),
            ]
        header += [
            text("column.compare.winner", locale),
            text("column.compare.limit", locale),
            text("column.compare.weight", locale),
            text("column.compare.note", locale),
        ]
        compare = sheet("heading.comparison")
        for column, label in enumerate(header, start=1):
            compare.cell(row=1, column=column, value=label).font = bold
        compare.freeze_panes = "A2"

        for index, row in enumerate(comparison, start=2):
            compare.cell(row=index, column=1, value=row.label)
            compare.cell(row=index, column=2, value=row.unit)
            column = 3
            for position, value in enumerate(row.values):
                cell = write_value(compare, index, column, value)
                # Only a candidate that really trails is marked. A cell
                # nobody measured did not lose the comparison — there was
                # no comparison — and a row where the two are level has no
                # loser either.
                if not value.missing and row.ahead:
                    cell.fill = ahead_fill if position in row.ahead else behind_fill
                if position in row.breaches:
                    cell.fill = breach_fill
                column += 1
            if has_delta:
                write_value(compare, index, column, row.delta or Quantity(None, COUNT))
                compare.cell(row=index, column=column + 1, value=row.delta_unit)
                column += 2
            compare.cell(row=index, column=column, value=row.winner)
            write_value(compare, index, column + 1, row.limit or Quantity(None, row.values[0].unit))
            write_value(compare, index, column + 2, row.weight or Quantity(None, WEIGHT))
            note = compare.cell(row=index, column=column + 3, value=row.note)
            note.alignment = wrap
            column += 4

        widths = [34, 8] + [16] * len(labels)
        if has_delta:
            widths += [12, 12]
        widths += [24, 14, 10, 80]
        for offset, width in enumerate(widths):
            compare.column_dimensions[get_column_letter(offset + 1)].width = width
        write_caveat(
            compare,
            text("heading.precision", locale),
            text("prose.summary_precision", locale),
            len(comparison) + 3,
        )

    # --- Objective breakdown ------------------------------------------------
    #
    # The only sheet where the weight column means anything in full.
    # Weights attach to objectives, not to the metrics a reader compares
    # on, which is why seven of the ten rows above carry none.
    objectives = objective_rows(run, report, locale, weights)
    if objectives:
        labels = [
            as_text(entry.get("stack_label"), locale) for entry in report.get("candidates") or []
        ]
        has_delta = any(row.delta is not None for row in objectives)
        header = [
            text("column.objective.name", locale),
            text("column.compare.weight", locale),
            *labels,
        ]
        if has_delta:
            header.append(text("column.compare.delta", locale))
        header += [
            f"{text('column.objective.contribution', locale)} — {label}" for label in labels
        ]
        header.append(text("column.compare.note", locale))

        axes = sheet("heading.objectives")
        for column, label in enumerate(header, start=1):
            axes.cell(row=1, column=column, value=label).font = bold
        axes.freeze_panes = "A2"

        # Eligibility leads, because it is the question the rest of the
        # sheet presumes an answer to. Text, not a number: yes/no is a
        # verdict, and a 1 or a 0 in a column of utilities would be
        # summed with them.
        eligible = eligibility_row(report, locale)
        offset = 0
        if eligible:
            label, verdicts = eligible
            axes.cell(row=2, column=1, value=label).font = bold
            for position, verdict in enumerate(verdicts):
                axes.cell(row=2, column=3 + position, value=verdict)
            offset = 1

        for index, row in enumerate(objectives, start=2 + offset):
            name = axes.cell(row=index, column=1, value=row.label)
            axes.cell(row=index, column=2).font = bold
            write_value(axes, index, 2, row.weight or Quantity(None, WEIGHT))
            column = 3
            for value in row.values:
                write_value(axes, index, column, value)
                column += 1
            if has_delta:
                write_value(axes, index, column, row.delta or Quantity(None, UTILITY))
                column += 1
            for value in row.contributions:
                cell = write_value(axes, index, column, value)
                if row.total:
                    cell.font = bold
                column += 1
            if row.note:
                note = axes.cell(row=index, column=column, value=row.note)
                note.alignment = wrap
            if row.total:
                name.font = bold

        widths = [30, 10] + [14] * len(labels)
        if has_delta:
            widths.append(12)
        widths += [20] * len(labels) + [80]
        for offset, width in enumerate(widths):
            axes.column_dimensions[get_column_letter(offset + 1)].width = width

        # Why a weight is missing, when one is. Three different absences
        # and only one of them is "this run had none".
        at = len(objectives) + offset + 3
        at = write_caveat(
            axes,
            text("column.outcome.eligible", locale),
            text("prose.eligible_sheet", locale),
            at,
        )
        if weights.reason:
            write_caveat(
                axes,
                text("heading.weights", locale),
                text(f"caveat.weights_{weights.reason}", locale, profile=weights.profile),
                at + 1,
            )

    # --- The recommendation, or why there is none -------------------------
    decision = sheet("heading.card")
    rows = card_rows(run, report, locale)
    if rows is None:
        decision.cell(
            row=1, column=1, value=text("heading.no_card", locale)
        ).font = bold
        reason = no_card_reason(report, locale)
        at = 3
        if reason:
            at = write_caveat(decision, text("heading.reason", locale), reason, at)
        write_caveat(
            decision,
            text("heading.what_this_means", locale),
            text("prose.no_card_sheet", locale),
            at,
        )
    else:
        cursor = write_pairs(decision, rows, 1)
        cursor = write_caveat(
            decision,
            text("heading.scope", locale),
            text("prose.scope_sheet", locale, scope=scope_of(run, report, locale)),
            cursor + 1,
        )
        evidence = decision_evidence_rows(run, report, locale)
        if evidence:
            decision.cell(
                row=cursor + 1, column=1, value=text("heading.margin", locale)
            ).font = bold
            cursor = write_pairs(decision, evidence, cursor + 2)
            cursor = write_caveat(
                decision,
                text("heading.reading_delta_u", locale),
                text("prose.delta_u_sheet", locale),
                cursor + 1,
            )

        card = run.card or report.get("decision_card") or {}
        margins = sensitivity_rows(card.get("evidence") or {}, locale)
        if margins is None:
            write_caveat(
                decision,
                text("heading.sensitivity", locale),
                text("prose.no_sensitivity_sheet", locale),
                cursor + 1,
            )
        else:
            decision.cell(
                row=cursor + 1, column=1, value=text("heading.sensitivity", locale)
            ).font = bold
            write_pairs(decision, margins, cursor + 2)
    decision.column_dimensions["A"].width = 26
    decision.column_dimensions["B"].width = 82

    # --- Every episode ----------------------------------------------------
    episodes = episode_rows(report, locale)
    if episodes:
        write_table(
            sheet("heading.episodes"), episode_columns(locale), episodes,
            (30, 20, 16, 11, 14, 12, 13, 10, 15),
        )

    # --- Human record -----------------------------------------------------
    human = sheet("heading.human")
    cursor = write_pairs(human, human_rows(run, locale), 1)
    write_caveat(
        human,
        text("heading.two_acts", locale),
        text("prose.two_acts_sheet", locale),
        cursor + 1,
    )
    human.column_dimensions["A"].width = 26
    human.column_dimensions["B"].width = 82

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
